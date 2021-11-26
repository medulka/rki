#!/usr/bin/env python3.8

"""
Reald World Data Evidence - Germany
DATE:   09.11.2021
(covid-19 day9 - covid-19 day14)
"""

from matplotlib import colors
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
from numpy.core.fromnumeric import size
from openpyxl.worksheet import worksheet
import requests
from pprint import pprint
import openpyxl
from datetime import datetime

#datove zdroje z webu www.rki.de
# url1 = 'https://raw.githubusercontent.com/robert-koch-institut/COVID-19-Impfungen_in_Deutschland/master/Aktuell_Deutschland_Bundeslaender_COVID-19-Impfungen.csv'
# url2 = 'https://raw.githubusercontent.com/robert-koch-institut/COVID-19-Hospitalisierungen_in_Deutschland/master/Aktuell_Deutschland_COVID-19-Hospitalisierungen.csv'

url3 = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Fallzahlen_Kum_Tab.xlsx?__blob=publicationFile"
url4 = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Inzidenz_Impfstatus.xlsx?__blob=publicationFile"
url5 = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquotenmonitoring.xlsx?__blob=publicationFile" 

#ulozeny xlsx dokument
fall = '/Users/hanamedova/Documents/COVID_TimeOutAcro/fall.xlsx'
impfquoten = '/Users/hanamedova/Documents/COVID_TimeOutAcro/impfquoten.xlsx'
fallzahlen_kum =  '/Users/hanamedova/Documents/COVID_TimeOutAcro/fallzahlen_kum.xlsx'

def xlsx_to_list(url, saved_xlsx_file, sheetname, minrow, n, m):
    "in: url for download, out: list; n - uprava vybranych radku, m - adjustace vybranych sloupcu"
    req = requests.get(url, allow_redirects=True)
    with open(saved_xlsx_file, 'wb') as f:
        f.write(req.content)
    workbook = openpyxl.load_workbook(saved_xlsx_file)
    worksheet = workbook[sheetname]
    maxrow = worksheet.max_row + n
    maxcolumn = worksheet.max_column + m
    out_list = [ cell.value for row in worksheet.iter_rows(min_row=minrow, max_row=maxrow, min_col=1, max_col=maxcolumn) for cell in row if cell.value != None ]
    return out_list

def main():
   
    #creating the data sets symptoms
    
    symptomatic_xlsx_list = xlsx_to_list(url=url4, saved_xlsx_file=fall, sheetname='Symptomatische_nach_Impfstatus', minrow=4, n=0, m=0)
    Meldewoche = symptomatic_xlsx_list[7::7] #preskocim pole "Meldewoche" 
    print('Meldewoche', Meldewoche, 'len(Meldewoche) ', len(Meldewoche))
    symptomatic_ungeimpfte_total = [ sum(x) for x in zip(symptomatic_xlsx_list[8::7], symptomatic_xlsx_list[10::7], symptomatic_xlsx_list[12::7]) ]
    symptomatic_geimpfte_total = [ sum(x) for x in zip(symptomatic_xlsx_list[9::7], symptomatic_xlsx_list[11::7], symptomatic_xlsx_list[13::7]) ]

    print('symptomatic_geimpfte_total ', symptomatic_geimpfte_total, 'len: ', len(symptomatic_geimpfte_total))

    #creating the data sets hospitalized

    hospitalized_xlsx_list = xlsx_to_list(url=url4, saved_xlsx_file=fall, sheetname='Hospitalisierte_nach_Impfstatus', minrow=4, n=0, m=0)
    Meldewoche_hosp = hospitalized_xlsx_list[8::7]
    print('len(Meldewoche_hosp): ', len(Meldewoche_hosp))
    hospitalized_ungeimpfte_total = [ sum(x) for x in zip(hospitalized_xlsx_list[8::7], hospitalized_xlsx_list[10::7], hospitalized_xlsx_list[12::7]) ]
    hospitalized_geimpfte_total = [ sum(x) for x in zip(hospitalized_xlsx_list[9::7], hospitalized_xlsx_list[11::7], hospitalized_xlsx_list[13::7]) ]


    #creating the data sets Impfquoten
    impfquoten_xlsx_list = xlsx_to_list(url=url5, saved_xlsx_file=impfquoten, sheetname='Impfungen_proTag', minrow=1, n=-6, m=0)
    Datum_interim_data = [datetime.strptime(i, '%d.%m.%Y') for i in impfquoten_xlsx_list[5::5]] #beginnt am Sonntag 27.12.2020 
    Datum = Datum_interim_data[ (7*29) : 7*(29+len(Meldewoche)) : 7 ] #beginnt am Sonntag 18.07.2021, endet am ande Meldewoche (44?)
    print('Datum: ', Datum, 'len(Datum): ', len(Datum) )


    Erste_Impfung_interim_data = np.cumsum(impfquoten_xlsx_list[6::5])
    Zweite_Impfung_interim_data = np.cumsum(impfquoten_xlsx_list[7::5])
    Dritte_Impfung_interim_data = np.cumsum(impfquoten_xlsx_list[8::5])
    
    #vybirame cast data setu odpovidajicimu Meldewochen
    Zweite_Impfung = Zweite_Impfung_interim_data[ (7*29) : 7*(29+len(Meldewoche)) : 7 ]
    Dritte_Impfung = Dritte_Impfung_interim_data[ (7*29) : 7*(29+len(Meldewoche)) : 7 ]
    print('len(Zweite_Impfung) ', len(Zweite_Impfung))
    print('len(Dritte_Impfung) ', len(Dritte_Impfung))

    #normalisiert fuer population / 100.000
    German_population_2020 = 83155031
    Zweite_Impfung_normalisiert = 100000 * Zweite_Impfung / German_population_2020
    Dritte_Impfung_normalisiert = 100000 * Dritte_Impfung / German_population_2020

    #data set Fallzahlen_kum
    fallzahlen_kum_xlsx = xlsx_to_list(url=url3, saved_xlsx_file=fallzahlen_kum, sheetname='BL_7-Tage-Inzidenz (fixiert)', minrow=3, n=0, m=0)
    len_row = int((len(fallzahlen_kum_xlsx)+1) / 18)
    Datum_fallzahlen_interim_data = fallzahlen_kum_xlsx[:(len_row-1)]
    #vybirame datum v souladu s delkou symptomatic / hospitaliziert / geimpft
    #1. den v datovem setu byla streda, o 62 tydnu vice nez v datovem setu symptoms
    Datum_fallzahlen = Datum_fallzahlen_interim_data[ (4 + 7*62) : 7 * (62 + len(Meldewoche)) : 7 ] 
    print('len(Datum_fallzahlen) ', len(Datum_fallzahlen))
    sieben_tage_inzidenz_interim_data = fallzahlen_kum_xlsx[-(len_row)+1:] # odectena erste Zell - 'Gesamt'
    sieben_tage_inzidenz = sieben_tage_inzidenz_interim_data[ (4 + 7*62) : 7 * (62 + len(Meldewoche)) : 7 ] 
    
    #quotienten 
    quotient_symp_unge_to_geimpt = [ a/b for a,b in zip(symptomatic_ungeimpfte_total, symptomatic_geimpfte_total) ]
    quotient_hosp_unge_to_geimpft = [ a/b for a,b in zip( hospitalized_ungeimpfte_total, hospitalized_geimpfte_total) ]
    quotient_hospit_to_symptom_ungeimpft = [ a/b for a,b in zip( hospitalized_ungeimpfte_total, symptomatic_ungeimpfte_total) ]
    quotient_hospit_to_symptom_geimpft = [ a/b for a,b in zip( hospitalized_geimpfte_total, symptomatic_geimpfte_total) ]

    q = [ a/b for a,b in zip( quotient_hospit_to_symptom_ungeimpft, quotient_hospit_to_symptom_geimpft) ] 
    r = [ b/a for a,b in zip( quotient_hospit_to_symptom_ungeimpft, quotient_hospit_to_symptom_geimpft) ] 

    #data for plotting
    # #A)
    # Datum
    # Zweite_Impfung_normalisiert
    # #Dritte_Impfung_normalisiert 
    
    # sieben_tage_inzidenz
    # symptomatic_geimpfte_total 
    # symptomatic_ungeimpfte_total

    # #B)
    # Datum
    # Zweite_Impfung_normalisiert
    # #Dritte_Impfung_normalisiert 
    # sieben_tage_inzidenz

    # hospitalized_ungeimpfte_total 
    # hospitalized_geimpfte_total 


    # #C)
    # Datum
    # Zweite_Impfung_normalisiert
    # #Dritte_Impfung_normalisiert 

    # quotient_symp_unge_to_geimpt 
    # quotient_hosp_unge_to_geimpft 
    
    # #D)   
    # Datum
    # Zweite_Impfung_normalisiert
    # #Dritte_Impfung_normalisiert 
    # quotient_hospit_to_symptom_ungeimpft 
    # quotient_hospit_to_symptom_geimpft 

    #plotting
    fig, ((ax1, ax3), (ax5, ax7)) = plt.subplots(2,2) 
    
    ax1.plot(Datum, sieben_tage_inzidenz, label = '7-day incidence')
    ax1.plot(Datum, symptomatic_geimpfte_total, label = 'symptomatic fully vaccinated')
    ax1.plot(Datum, symptomatic_ungeimpfte_total, label = 'symptomatic unvaccinated')
    
    ax1.set_title('Symptomatic cases per 100.000', color = 'b')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Cases')
    ax1.legend(loc = 'upper right')
    ax1.grid()

    ax2 = ax1.twinx()
    ax2.fill_between(Datum, Zweite_Impfung_normalisiert, Zweite_Impfung_normalisiert.min(), alpha = 0.3, label = 'Doses registered')
    ax2.set_ylabel('Fully vaccinated')
    
    #druhy graf

    ax3.plot(Datum, sieben_tage_inzidenz, label = '7-day incidence')
    ax3.plot(Datum, hospitalized_ungeimpfte_total, label = 'hospotalized unvaccinated')
    ax3.plot(Datum, hospitalized_geimpfte_total, label = 'hospitalized fully vaccinated')
    
    ax4 = ax3.twinx()
    ax4.fill_between(Datum, Zweite_Impfung_normalisiert, Zweite_Impfung_normalisiert.min(), alpha = 0.3, label = 'Doses registered')
    ax4.set_ylabel('Fully vaccinated')

    ax3.legend(loc = 'upper right')

    #treti graf
    
    ax5.plot(Datum, quotient_symp_unge_to_geimpt, label = 'Unge_to_geimpt')
    ax5.plot(Datum, quotient_hosp_unge_to_geimpft, label = 'quotient_hosp_unge_to_geimpft')
    
    ax6 = ax5.twinx()
    ax6.fill_between(Datum, Zweite_Impfung_normalisiert, Zweite_Impfung_normalisiert.min(), alpha = 0.3, label = 'Doses registered')
    ax6.set_ylabel('Fully vaccinated')

    ax5.legend(loc = 'upper right')

    #ctvrty graf
    
    ax7.plot(Datum, quotient_hospit_to_symptom_ungeimpft, label = 'quotient_hospit_to_symptom_ungeimpft')
    ax7.plot(Datum, quotient_hospit_to_symptom_geimpft, label = 'quotient_hospit_to_symptom_geimpft')

    ax8 = ax7.twinx()
    ax8.fill_between(Datum, Zweite_Impfung_normalisiert, Zweite_Impfung_normalisiert.min(), alpha = 0.3, label = 'Doses registered')
    ax8.set_ylabel('Fully vaccinated')

    ax7.legend(loc = 'upper right')

    fig.tight_layout()
    plt.show()
    
    #title 'Inzidenzen der symptomatischen und hospitalisierten COVID-19-Fälle nach Impfstatus'

if __name__ == '__main__':
    main()


#what I leart meanwhile:
# A) with closing(requests.get(url1b, stream=True)) as r:
#         reader = csv.reader(r.iter_lines(), delimiter = ',')  #iteration over the lines, improve memory performance
#         for row in reader:
#             print(row)

#B) error_bad_lines might be useful
    # download = requests.get(url1).content #click on raw file in GitHub
    # df = pd.read_csv(io.StringIO(download.decode('utf-8')), error_bad_lines=False)
    # print(df)


#from urllib.request import Request, urlopen
#when it returns error 403 Forbidden - use headers to explicilty specified the user aggents
# headers ={'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'}
#     req = Request(url3, headers)
#     file = urlopen(req).read()
#file_decoded = file.decode('ISO-8859-1')  #or encoding='ISO-8859-1'
#    req = requests.get(url4, allow_redirects=True)
#     print(req.url)
#     req.text
#     req.encoding
#     req.content
#     req.raw
#     req.raw.read()
#     #url3_decoded = file.decode('ISO-8859-1')

#     with open('fall.xlsx', 'wb') as fl:
#         for chunk in req.iter_content():
#             fl.write(chunk)
#             print(chunk)
#     print(fl)

#www.destatis.de
#https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/Bevoelkerungsstand/Tabellen/bevoelkerung-altersgruppen-deutschland.html
#Stand 2020: 83155031

# file = pd.read_csv(url1, error_bad_lines=False)
# pprint(file)

#notes: all data are normalized per 100.000 population, 
#the population of 83155031 was considered (www.destatis.de, status: 2020)
#the data source: www.rki.de
#the time window is restricted by available data of (un-/)vaccinated, (un-/)symptomatic, hospitalized cases
#only data for "Meldewoche 28 - 44", i.e. 18.07.2021 - 07.11.2021 are available. For an overview of recent burning situation you will have to wait till Christmas. 
#the data set 7-day inzidenz and symptomatic cases should not be considered as corresponding, homogenous. 
#Effectiveness of vaccination campaings in Germany - Real World Data Evidence
#mind another scaling of y-axis



